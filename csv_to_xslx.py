import os
import csv
from openpyxl import Workbook
from pathlib import Path


def csv_to_excel(folder_path: Path, output_file):
    """
    指定されたフォルダ内のCSVをそれぞれのシートにして1つのExcelファイルにまとめる。

    Args:
        folder_path (str): CSVファイルが格納されたフォルダのパス
        output_file (str): 作成されるExcelファイルのパス
    """
    # 新しいExcelワークブックを作成
    wb = Workbook()
    initial_sheet = wb.active
    initial_sheet.title = "概要"  # 最初のシートを概要用にする

    # フォルダ内のCSVファイルを取得
    try:
        csv_files = [f for f in os.listdir(folder_path) if f.endswith(".csv")]
    except FileNotFoundError as e:
        print("フォルダが存在しません、以下から指定してください。")
        print(os.listdir(folder_path.parent))
        exit(print(e))

    if not csv_files:
        print("指定されたフォルダにCSVファイルがありません。")
        return

    for csv_file in csv_files:
        csv_path = os.path.join(folder_path, csv_file)

        # CSVを新しいシートに追加
        sheet_name = os.path.splitext(csv_file)[
            0
        ]  # 拡張子を除いたファイル名をシート名にする
        ws = wb.create_sheet(title=sheet_name)

        with open(csv_path, mode="r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                ws.append(row)

    # 最初の空シート（"概要"）を削除
    if "概要" in wb.sheetnames and len(wb.sheetnames) > 1:
        wb.remove(initial_sheet)

    # Excelファイルを保存
    wb.save(output_file)
    print(f"Excelファイルが作成されました：{output_file}")


# フォルダパスと出力先ファイル名を指定
folder_path = Path("./csv/" + input())  # CSVファイルが格納されたフォルダのパス
output_file = folder_path / "merged_data.xlsx"  # 出力されるExcelファイル名

csv_to_excel(folder_path, output_file)
