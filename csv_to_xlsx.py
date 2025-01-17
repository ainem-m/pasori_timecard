import csv
import json
from openpyxl import load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill
from openpyxl.worksheet.cell_range import CellRange
from pathlib import Path
import config  # config.pyにEMPLOYEE_LISTが設定されている前提
import to_csv
import time_util
from openpyxl.styles import Font, PatternFill, Border, Alignment
import sys

PRINT_AREA = "O1:V38"
VALIDATE_AREA = CellRange("G2:G33")
HIGHLIGHT_AREA = CellRange("A2:G33")


def load_employee_mapping(file_path: Path) -> dict:
    """
    従業員名とテンプレート種別のマッピング情報を読み込む。

    Args:
        file_path (Path): JSON形式のマッピングファイルパス

    Returns:
        dict: 従業員名とテンプレート種別のマッピング
    """
    try:
        with file_path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        print(f"{file_path} が見つかりません。")
        exit()
    except json.JSONDecodeError:
        print(f"{file_path} の形式が正しくありません。")
        exit()


def get_csv_files(folder_path: Path) -> list:
    """
    指定されたフォルダ内のCSVファイルを取得。

    Args:
        folder_path (Path): CSVファイルが格納されたフォルダのパス

    Returns:
        list: CSVファイルPathのリスト
    """
    if not folder_path.exists():
        print(f"フォルダ {folder_path} が存在しません。")
        exit()

    return [file for file in folder_path.iterdir() if file.suffix == ".csv"]


def write_csv_to_sheet(ws, csv_path: Path):
    """
    CSVの内容を指定されたシートに書き込む。

    Args:
        ws (Worksheet): 書き込み先のシート
        csv_path (Path): CSVファイルのパス
    """
    with csv_path.open(mode="r", encoding="utf-8") as f:
        reader = csv.reader(f)
        for row_idx, row in enumerate(reader, start=1):  # 行番号は1から始まる
            for col_idx, value in enumerate(row, start=1):  # 列番号は1から始まる
                cell_value = ws[row_idx][col_idx].value
                if cell_value in [None, "", to_csv.BLANK, to_csv.LOST]:
                    # 今まで入力されたことのないセル、もしくはLOSTのセルのみ入力する
                    ws.cell(row=row_idx, column=col_idx, value=value)


def copy_sheet(source_ws, target_wb, target_name):
    """
    wb1のws1シートをwb2にコピーする。値、書式、列の幅、行の高さを含む。

    Args:
        source_ws (Worksheet): コピー元のシート
        target_wb (Workbook): コピー先のワークブック
        target_name (str): コピー先のシート名

    Returns:
        bool: コピーに成功すればTrue,すでに存在していた場合False
    """
    # 新しいシートを作成
    if target_name in target_wb.sheetnames:
        return False
    ws2 = target_wb.create_sheet(target_name)

    # セルの値と書式をコピー
    for row in source_ws.iter_rows():
        for cell in row:
            # 新しいシートにセルの値をコピー
            new_cell = ws2[cell.coordinate]
            new_cell.value = cell.value

            # スタイル（フォント、色、罫線、配置など）をコピー
            if cell.font:
                new_cell.font = Font(
                    name=cell.font.name,
                    size=cell.font.size,
                    bold=cell.font.bold,
                    italic=cell.font.italic,
                    vertAlign=cell.font.vertAlign,
                    underline=cell.font.underline,
                    strike=cell.font.strike,
                    color=cell.font.color,
                )

            if cell.fill:
                new_cell.fill = PatternFill(
                    patternType=cell.fill.patternType,
                    fgColor=cell.fill.fgColor,
                    bgColor=cell.fill.bgColor,
                )

            if cell.border:
                new_cell.border = Border(
                    left=cell.border.left,
                    right=cell.border.right,
                    top=cell.border.top,
                    bottom=cell.border.bottom,
                    diagonal=cell.border.diagonal,
                    outline=cell.border.outline,
                    vertical=cell.border.vertical,
                    horizontal=cell.border.horizontal,
                )

            if cell.alignment:
                new_cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    textRotation=cell.alignment.textRotation,
                    wrapText=cell.alignment.wrapText,
                    shrinkToFit=cell.alignment.shrinkToFit,
                    indent=cell.alignment.indent,
                    relativeIndent=cell.alignment.relativeIndent,
                )

            if cell.number_format:
                new_cell.number_format = cell.number_format

    # 列幅をコピー
    for col_letter, column_dimension in source_ws.column_dimensions.items():
        ws2.column_dimensions[col_letter].width = column_dimension.width

    # 行の高さをコピー
    for row_number, row_dimension in source_ws.row_dimensions.items():
        ws2.row_dimensions[row_number].height = row_dimension.height

    return True


def initialize_sheet(ws) -> None:
    """
    シートの印刷範囲（PRINT_AREA）を設定し、「備考」欄（VALIDATE_AREA）に入力規則を適用する
    """
    # 印刷範囲の設定
    ws.print_area = PRINT_AREA

    # 備考欄の入力規則
    validation = DataValidation(
        type="list",
        formula1='"有給, AM有給, PM有給, 早退, 遅刻, 欠勤, 中抜け, 追加残業"',
        allow_blank=True,
    )
    validation.error = "選択肢から選んでください"
    validation.errorTitle = "入力エラー"
    validation.prompt = "選択肢から選んでください"
    validation.promptTitle = "備考欄の入力"
    # シートに入力規則を追加
    ws.add_data_validation(validation)
    validation.add(VALIDATE_AREA)

    # 条件付き書式の追加 なぜかうまくいかない
    # red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
    # rule = FormulaRule(formula=['$F2="-"'], stopIfTrue=True, fill=red_fill)
    # ws.conditional_formatting.add(HIGHLIGHT_AREA, rule)


def csv_to_excel(
    folder_path: Path, template_file: Path, employee_mapping: dict, output_file: Path
):
    """
    CSVデータを従業員名に対応するテンプレートシートに書き込む。

    Args:
        folder_path (Path): CSVファイルが格納されたフォルダのパス
        template_file (Path): テンプレートExcelファイルのパス
        employee_mapping (dict): 従業員名とテンプレートのマッピング
        output_file (Path): 出力されるExcelファイルのパス
    """
    year = int(folder_path.name.split("-")[0])
    month = int(folder_path.name.split("-")[1])

    # テンプレートファイルを読み込む
    if not template_file.exists():
        print(f"テンプレートファイル {template_file} が存在しません。")
        return

    template = load_workbook(template_file, keep_vba=True)
    temp_types = template.sheetnames
    if not output_file.exists():
        template.save(output_file)

    wb = load_workbook(output_file, keep_vba=True)
    wb.defined_names["year"] = DefinedName(name="year", attr_text=year)
    wb.defined_names["month"] = DefinedName(name="month", attr_text=month)

    # CSVファイルを取得
    csv_files = get_csv_files(folder_path)
    if not csv_files:
        print("指定されたフォルダにCSVファイルがありません。")
        return

    # 各CSVファイルを処理
    for csv_file in csv_files:
        employee_name = csv_file.stem
        template_type = employee_mapping.get(employee_name)

        if not template_type:
            print(
                f"{employee_name} のテンプレート情報が見つかりません。スキップします。"
            )
            continue

        # テンプレートを従業員名のシートとしてコピー
        if copy_sheet(template[template_type], wb, employee_name):
            initialize_sheet(wb[employee_name])
        else:
            print(employee_name, "がすでに存在していたため上書きします")

        # CSVデータを書き込み
        write_csv_to_sheet(wb[employee_name], csv_file)

    for temp in temp_types:
        if temp == "button":
            continue
        # 初期Sheetを削除
        if temp in wb.sheetnames:
            del wb[temp]

    # 新しいExcelファイルを保存
    wb.save(output_file)
    print(f"新しいExcelファイルが作成されました：{output_file}")


# メイン処理
if __name__ == "__main__":
    input_str: str
    if len(sys.argv) != 2:
        input_str = input("年と月を入力 例: 2024/08, 2024/8, 24/08, 24/8 ->")
    else:
        input_str = sys.argv[1]
    year, month = time_util.parse_date_string(input_str)
    folder_path = Path(config.CSV_PATH) / f"{year:04d}-{month:02d}"
    # CSVファイルが格納されたフォルダのパス
    template_file = Path(config.TEMPLATE_PATH)  # テンプレートExcelファイルのパス
    output_file = folder_path / "data.xlsm"  # 出力される新規Excelファイルのパス

    # 従業員名とテンプレートのマッピングを読み込む
    employee_mapping = load_employee_mapping(Path(config.EMPLOYEE_LIST))

    # CSVからExcelへのデータ転送
    csv_to_excel(folder_path, template_file, employee_mapping, output_file)
